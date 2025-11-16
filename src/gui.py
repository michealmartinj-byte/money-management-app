"""Beautiful Tkinter GUI for the Money Management App (percentage-based Martingale).

Run with:
  python -m src.gui

Modern dark theme with enhanced UI/UX, status messages (no popups), and Excel export.
"""
import tkinter as tk
from tkinter import ttk
import tkinter.font as tkFont
from money_manager.session import Account
import os
from pathlib import Path

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Color scheme (modern dark theme)
DARK_BG = "#1e1e1e"
PRIMARY_COLOR = "#2ecc71"  # Green
SECONDARY_COLOR = "#3498db"  # Blue
ACCENT_COLOR = "#e74c3c"  # Red (for losses)
TEXT_COLOR = "#ecf0f1"  # Light gray
HEADER_BG = "#2c3e50"  # Dark blue-gray
SUCCESS_COLOR = "#27ae60"  # Darker green
WARNING_COLOR = "#e67e22"  # Orange


class MoneyApp(tk.Tk):
    def __init__(self, data_file=None):
        super().__init__()
        self.title("💰 Money Manager - Martingale Trading")
        self.geometry("950x800")
        self.configure(bg=DARK_BG)
        self.resizable(True, True)

        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TFrame', background=DARK_BG)
        style.configure('TLabel', background=DARK_BG, foreground=TEXT_COLOR)
        style.configure('TButton', background=PRIMARY_COLOR, foreground='black')
        style.configure('TLabelframe', background=DARK_BG, foreground=TEXT_COLOR)
        style.configure('Treeview', background='#2c3e50', foreground=TEXT_COLOR, fieldbackground='#2c3e50', borderwidth=0)
        style.configure('Treeview.Heading', background=HEADER_BG, foreground=TEXT_COLOR)

        self.account = Account(data_file=data_file)
        self.excel_file = os.path.join(os.path.dirname(data_file) if data_file else os.getcwd(), "trading_history.xlsx")

        # Main container
        main_container = tk.Frame(self, bg=DARK_BG)
        main_container.pack(fill=tk.BOTH, expand=True, padx=16, pady=16)

        # Title
        title_font = tkFont.Font(family="Segoe UI", size=18, weight="bold")
        title = tk.Label(main_container, text="💰 Money Manager", font=title_font, bg=DARK_BG, fg=PRIMARY_COLOR)
        title.pack(anchor=tk.W, pady=(0, 16))

        # Status bar (replaces popups)
        status_frame = tk.Frame(main_container, bg=HEADER_BG, relief=tk.FLAT, bd=1, height=30)
        status_frame.pack(fill=tk.X, pady=(0, 12))
        status_frame.pack_propagate(False)
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = tk.Label(status_frame, textvariable=self.status_var, bg=HEADER_BG, fg=SUCCESS_COLOR, font=("Segoe UI", 9))
        self.status_label.pack(anchor=tk.W, padx=12, pady=6)

        # Top section: Balance & Init
        top_frame = tk.Frame(main_container, bg=HEADER_BG, relief=tk.FLAT, bd=1)
        top_frame.pack(fill=tk.X, pady=(0, 12))

        bal_label = tk.Label(top_frame, text="Current Balance", font=("Segoe UI", 10), bg=HEADER_BG, fg=TEXT_COLOR)
        bal_label.pack(anchor=tk.W, padx=12, pady=(8, 2))

        self.balance_var = tk.StringVar(value=str(self.account.balance))
        balance_font = tkFont.Font(family="Segoe UI", size=28, weight="bold")
        self.balance_lbl = tk.Label(top_frame, textvariable=self.balance_var, font=balance_font, bg=HEADER_BG, fg=PRIMARY_COLOR)
        self.balance_lbl.pack(anchor=tk.W, padx=12, pady=(0, 8))

        init_row = tk.Frame(top_frame, bg=HEADER_BG)
        init_row.pack(fill=tk.X, padx=12, pady=(0, 12))
        tk.Label(init_row, text="Init Balance:", bg=HEADER_BG, fg=TEXT_COLOR, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.init_entry = tk.Entry(init_row, width=12, bg='#3c3f41', fg=TEXT_COLOR, insertbackground=TEXT_COLOR, font=("Segoe UI", 10), relief=tk.FLAT, bd=1)
        self.init_entry.pack(side=tk.LEFT, padx=(4, 8))
        btn_init = tk.Button(init_row, text="Initialize", command=self.on_init, bg=PRIMARY_COLOR, fg='black', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=4, cursor="hand2")
        btn_init.pack(side=tk.LEFT)

        # Session & Betting section
        mid_frame = tk.Frame(main_container, bg=HEADER_BG, relief=tk.FLAT, bd=1)
        mid_frame.pack(fill=tk.X, pady=(0, 12))

        session_label = tk.Label(mid_frame, text="Session & Betting", font=("Segoe UI", 11, "bold"), bg=HEADER_BG, fg=PRIMARY_COLOR)
        session_label.pack(anchor=tk.W, padx=12, pady=(8, 8))

        btn_row1 = tk.Frame(mid_frame, bg=HEADER_BG)
        btn_row1.pack(fill=tk.X, padx=12, pady=(0, 8))
        tk.Button(btn_row1, text="▶ Start Session", command=self.on_start_session, bg=SECONDARY_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_row1, text="⏹ End Session", command=self.on_end_session, bg=ACCENT_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=6)

        bet_row = tk.Frame(mid_frame, bg=HEADER_BG)
        bet_row.pack(fill=tk.X, padx=12, pady=(0, 8))
        tk.Label(bet_row, text="Base %:", bg=HEADER_BG, fg=TEXT_COLOR, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.base_percent_var = tk.DoubleVar(value=0.02)
        tk.Entry(bet_row, textvariable=self.base_percent_var, width=6, bg='#3c3f41', fg=TEXT_COLOR, insertbackground=TEXT_COLOR, font=("Segoe UI", 9), relief=tk.FLAT, bd=1).pack(side=tk.LEFT, padx=(4, 16))
        tk.Label(bet_row, text="Multiplier:", bg=HEADER_BG, fg=TEXT_COLOR, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.mult_var = tk.DoubleVar(value=2.0)
        tk.Entry(bet_row, textvariable=self.mult_var, width=6, bg='#3c3f41', fg=TEXT_COLOR, insertbackground=TEXT_COLOR, font=("Segoe UI", 9), relief=tk.FLAT, bd=1).pack(side=tk.LEFT, padx=(4, 16))
        tk.Button(bet_row, text="📊 Next Bet", command=self.on_next_bet, bg=SECONDARY_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=4, cursor="hand2").pack(side=tk.LEFT, padx=(0, 8))
        self.next_bet_var = tk.StringVar(value="-")
        tk.Label(bet_row, textvariable=self.next_bet_var, bg=HEADER_BG, fg=PRIMARY_COLOR, font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT)

        btn_row2 = tk.Frame(mid_frame, bg=HEADER_BG)
        btn_row2.pack(fill=tk.X, padx=12, pady=(0, 12))
        tk.Button(btn_row2, text="✓ Record Win", command=self.on_record_win, bg=PRIMARY_COLOR, fg='black', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_row2, text="✗ Record Loss", command=self.on_record_loss, bg=ACCENT_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=6)
        tk.Button(btn_row2, text="💾 Export to Excel", command=self.export_to_excel, bg=WARNING_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=6)

        # Session history
        hist_label = tk.Label(main_container, text="Current Session History", font=("Segoe UI", 11, "bold"), bg=DARK_BG, fg=PRIMARY_COLOR)
        hist_label.pack(anchor=tk.W, pady=(12, 4))

        hist_frame = tk.Frame(main_container, bg=HEADER_BG, relief=tk.FLAT, bd=1)
        hist_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 12))

        columns = ("step", "bet_amount", "bet_pct", "result", "pnl", "balance")
        column_names = {"step": "Step", "bet_amount": "Bet $ Amount", "bet_pct": "Bet %", "result": "Win/Loss", "pnl": "P&L ($)", "balance": "New Balance"}
        self.tree = ttk.Treeview(hist_frame, columns=columns, show="headings", height=10)
        for c in columns:
            self.tree.heading(c, text=column_names[c])
            self.tree.column(c, width=135)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Configure treeview colors for wins/losses
        self.tree.tag_configure('win', foreground=PRIMARY_COLOR)
        self.tree.tag_configure('loss', foreground=ACCENT_COLOR)

        # Footer
        footer_frame = tk.Frame(main_container, bg=DARK_BG)
        footer_frame.pack(fill=tk.X, pady=(0, 0))
        tk.Button(footer_frame, text="🔄 Refresh", command=self.refresh, bg=SECONDARY_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=4, cursor="hand2").pack(side=tk.LEFT)
        tk.Button(footer_frame, text="❌ Quit", command=self.quit, bg=ACCENT_COLOR, fg='white', font=("Segoe UI", 9, "bold"), relief=tk.FLAT, bd=0, padx=12, pady=4, cursor="hand2").pack(side=tk.RIGHT)

        self.refresh()

    def set_status(self, message, color=SUCCESS_COLOR):
        """Update status bar instead of showing popups"""
        self.status_var.set(message)
        self.status_label.config(fg=color)
        self.after(3000, lambda: self.status_var.set("Ready"))  # Reset after 3 seconds

    def on_init(self):
        try:
            val = float(self.init_entry.get())
        except Exception:
            self.set_status("❌ Invalid balance - enter a number", ACCENT_COLOR)
            return
        self.account.init_account(val)
        self.init_entry.delete(0, tk.END)
        self.refresh()
        self.export_to_excel()  # Auto-export after init
        self.set_status(f"✓ Account initialized with balance: {val}", SUCCESS_COLOR)

    def on_start_session(self):
        try:
            sess = self.account.start_session()
            self.set_status(f"✓ Session started at {sess.start_balance}", SUCCESS_COLOR)
        except Exception as e:
            self.set_status(f"⚠ {str(e)}", WARNING_COLOR)
        self.refresh()

    def on_end_session(self):
        self.account.force_end_session()
        self.set_status("✓ Session ended", SUCCESS_COLOR)
        self.refresh()

    def on_next_bet(self):
        info = self.account.get_next_bet(base_percent=self.base_percent_var.get(), multiplier=self.mult_var.get())
        self.next_bet_var.set(f"{info['bet_amount']:.2f} ({info['bet_percent']*100:.2f}%)")
        self.set_status(f"Next bet: ${info['bet_amount']:.2f} ({info['bet_percent']*100:.2f}%)", SECONDARY_COLOR)

    def _record(self, win: bool):
        if self.account.current_session is None:
            try:
                self.account.start_session()
            except Exception:
                pass

        info = self.account.get_next_bet(base_percent=self.base_percent_var.get(), multiplier=self.mult_var.get())
        bet_percent = info["bet_percent"]
        bet_amount = info["bet_amount"]

        if win:
            pnl = float(bet_amount)
        else:
            pnl = -float(bet_amount)

        try:
            step = self.account.record_result(bet_percent=bet_percent, bet_amount=bet_amount, pnl=pnl, result=("win" if win else "loss"))
            result_text = "WIN ✓" if win else "LOSS ✗"
            self.set_status(f"Step {step.idx}: {result_text} | P&L: {step.pnl:+.2f} | Balance: {step.balance_after:.2f}", SUCCESS_COLOR if win else ACCENT_COLOR)
            self.export_to_excel()  # Auto-export after each trade
        except Exception as e:
            self.set_status(f"❌ Error: {str(e)}", ACCENT_COLOR)

        self.refresh()

    def on_record_win(self):
        self._record(True)

    def on_record_loss(self):
        self._record(False)

    def export_to_excel(self):
        """Export all sessions to Excel file"""
        if not HAS_OPENPYXL:
            self.set_status("⚠ openpyxl not installed - install with: pip install openpyxl", WARNING_COLOR)
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Trading History"

            # Headers
            headers = ["Step", "Bet Amount ($)", "Bet %", "Win/Loss", "P&L ($)", "New Balance"]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_font = Font(bold=True, color="ecf0f1")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add all sessions data
            row_num = 2
            for session in self.account.sessions:
                for step in session.steps:
                    ws.append([
                        step.idx,
                        f"{step.bet_amount:.2f}",
                        f"{step.bet_percent*100:.2f}%",
                        step.result.upper(),
                        f"{step.pnl:+.2f}",
                        f"{step.balance_after:.2f}"
                    ])
                    # Color code rows
                    if step.result == "win":
                        row_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
                    else:
                        row_fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")
                    for cell in ws[row_num]:
                        cell.fill = row_fill
                        cell.font = Font(color="ecf0f1")
                        cell.alignment = Alignment(horizontal="center")
                    row_num += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15

            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["Metric", "Value"])
            summary_ws.append(["Starting Balance", ""])
            summary_ws.append(["Current Balance", self.account.balance])
            summary_ws.append(["Total Profit/Loss", self.account.balance - (self.account.sessions[0].start_balance if self.account.sessions else self.account.balance)])
            summary_ws.append(["Total Sessions", len(self.account.sessions)])
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 15

            wb.save(self.excel_file)
            self.set_status(f"✓ Exported to Excel: {Path(self.excel_file).name}", SUCCESS_COLOR)
        except Exception as e:
            self.set_status(f"❌ Excel export failed: {str(e)}", ACCENT_COLOR)

    def refresh(self):
        self.balance_var.set(f"{self.account.balance:.2f}")
        
        # Update balance color
        if self.account.balance > 1000:
            self.balance_lbl.config(fg=PRIMARY_COLOR)
        elif self.account.balance < 1000:
            self.balance_lbl.config(fg=ACCENT_COLOR)
        else:
            self.balance_lbl.config(fg=PRIMARY_COLOR)

        # Update session tree
        for i in self.tree.get_children():
            self.tree.delete(i)

        sess = self.account.current_session
        if sess:
            for st in sess.steps:
                tag = 'win' if st.result == 'win' else 'loss'
                self.tree.insert("", tk.END, values=(st.idx, f"{st.bet_amount:.2f}", f"{st.bet_percent*100:.2f}%", st.result, f"{st.pnl:.2f}", f"{st.balance_after:.2f}"), tags=(tag,))


def main():
    app = MoneyApp()
    app.mainloop()


if __name__ == "__main__":
    main()
